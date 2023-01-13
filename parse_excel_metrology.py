"""
Модуль для парсинга файла Excel с перечнём средств измерений
и метрологией по ним.
"""
import os
import sys
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill
import argparse
import re
import app_logger
import parse_fgis
from work_db import WorkDb
from parse_type_si import TypeParseSi

logger = app_logger.get_logger(__name__)
# Константы
# ================== #

# Номера столбцов для соответствующего типа СИ
COLUMNS_SI = {
    'ПУ': {
        'type': 8,
        'number': 9,
        'verif_date': 12,
        'valid_date': 13,
        'href': 14
    },
    'ТТ': {
        'type': 15,
        'number': 17,
        'verif_date': 20,
        'valid_date': 21,
        'href': 22
    },
    'ТН': {
        'type': 23,
        'number': 25,
        'verif_date': 28,
        'valid_date': 29,
        'href': 30
    }
}

EXC_STR = ['',
           '-',
           '--',
           '---',
           'не пригоден',
           'н/д',
           'нет данных',
           'отсутствует',
           None]

# Начальная строка
START_ROW = 13

# =================== #


def create_parse_arg():
    """
    Парсер для параметров командной строки,
    разбор переданных параметров запуска
    """
    logger.info("Парсинг параметров командной строки")
    parser = argparse.ArgumentParser()
    parser.add_argument('-s', '--serial', type=str, default='', help='Конкретный номер СИ для поиска в БД FGIS')
    parser.add_argument('-n', '--namefile', type=str, default="Перечень_ТУ_АСКУЭ_ООСС_АСКУЭ_4_кв_2022.xlsx",
                        help='Имя файла Excel')
    parser.add_argument('-y', '--years', type=int, default=datetime.today().year,
                        help='Год поверки СИ для выборки')
    parser.add_argument('-k', '--keyword', type=str, default='ТТ', help="СИ по которому нужно получить данные из ФГИС: "
                                                                              "ПУ - приборы учёта, ТТ - трансформаторы тока, ТН - трансформаторы напряжения.")
    parser.add_argument('-m', '--mode', type=str, default='fgis', help="Режим запуска скрипта: fgis - для сбора данных из БД ФГИС,"
                                                                        "local - для локальной работы и заполнения файла Excel,"
                                                                       "change_serial - для добавления 0 вначале номера для ЕвроАльфа")
    logger.info("Парсинг параметров командной строки закончен")

    return parser


def get_worksheet(namefile: str):
    """
    Открываем файл Excel, возвращаем объект книги и листа

    :param namefile: имя файла
    :return: workbook, worksheet - объект книги и объект листа
    """
    logger.info(f"Пытаемся открыть файл {namefile}")
    if os.path.isfile(namefile):
        workbook = openpyxl.load_workbook(namefile, data_only=True)
        worksheet = workbook['Прил.1.1 (Сч,ТТ,ТН)']
        logger.info(f"Файл {namefile} успешно открыт")
        return workbook, worksheet
    else:
        logger.warning(f"Невозможно открыть файл {namefile}, так как он не существует")


def get_verif_year_from_str(str_verif_year):
    """
    Метод для получения года поверки из строки

    :param str_verif_year: строка - дата последние поверки
    :return: verif_year: int - год последней поверки
    """

    try:
        logger.info(f"Попытка получить из строки {str_verif_year} год")
        if type(str_verif_year) == datetime:
            verif_year = str_verif_year.date().year
        elif type(str_verif_year) == str:
            tmp_datetime = datetime.strptime(str_verif_year, "&d.%m.%Y")
            verif_year = tmp_datetime.date().year
        print(f"Год поверки для текущего СИ - {verif_year}")
        logger.info(f"Год поверки для текущего СИ - {verif_year}")
        return verif_year
    except:
        logger.warning(f"Не удалось получить год поверки из строки - {str_verif_year}")


def format_dict_requests(title="", verif_year=datetime.today().year, mitype=None, number="", rows=str(20)):
    """
    Формируем словарь для формирования запроса в БД ФГИС

    :param title: наименование типа СИ, часть или полное наименование
    :param verif_year: год последней поверки, для выборки из ФГИС
    :param number: заводской номер СИ для выборки из ФГИС
    :param rows: количество строк в запросе
    :return: dict_requests - сформированный словарь для запроса
    """
    if title == "ПУ":
        filter_mititle = "Счётчик*электрической*%20"
    elif title == "ТТ":
        filter_mititle = "Трансформаторы*тока*%20"
    elif title == "ТН":
        filter_mititle = "Трансформаторы*напряжения*%20"

    dict_request = {
        'filter_mititle': filter_mititle,
        'verification_year': str(verif_year),
        'filter_minumber': str(number),
        'filter_mitype': mitype,
        'rows': rows
    }

    return dict_request

def change_serial_number(serial: str):
    pass

def get_parse_si(si_for_parse: str, verif_year, mode, namefile='', serial=''):
    """
    Парсинг из файла средства измерений, переданного в строке si_for_parse

    :param si_for_parse: СИ для парсинга
    :param verif_year: год поверки дл выборки
    :param mode: режим работы: fgis - для сбора данных из БД ФГИС по СИ и запись в локальную БД
                для дальнейшей выборки оттуда гиперссылок;
                               local - для работы с локальной БД, и заполнения информации по СИ в файле
    :param namefile: имя Excel файла для работы, по умолчанию из файла берётся конкретный лист, с конкретным названием
    :return:
    """
    logger.info("Старт get_parse_si()")
    Fill = PatternFill(fill_type='solid', start_color='00FA9A', end_color='00FA9A')
    FillRed = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    FillOrange = PatternFill(fill_type='solid', start_color='FF66FF', end_color='FF66FF')
    FillBlue = PatternFill(fill_type='solid', start_color='0099FF', end_color='0099FF')
    FillYellow = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')

    # Открываем файл и получаем объект листа
    workbook, worksheet = get_worksheet(namefile)
    if workbook != None and worksheet != None:
        end_row = worksheet.max_row

        # В цикле проходим по всем СИ
        for si in si_for_parse.split(sep=" "):
            # Получаем номера столбцов для соответствующего СИ
            # из словаря констант
            type_col = COLUMNS_SI[si]['type']
            number_col = COLUMNS_SI[si]['number']
            verif_date_col = COLUMNS_SI[si]['verif_date']
            valid_date_col = COLUMNS_SI[si]['valid_date']
            href_col = COLUMNS_SI[si]['href']
            logger.info(f"Старт обхода строк файла {namefile} для СИ: {', '.join(si_for_parse.split(sep=' '))}")
            for r in range(START_ROW, end_row + 1):
                if si == "ПУ":
                    serial = str(worksheet.cell(row=r, column=number_col).value)
                    if serial[:2] == "12":
                        worksheet.cell(row=r, column=number_col).value = '0' + serial
                print(f"Строка №{r} файла Excel.")
                str_verif_date = worksheet.cell(row=r, column=verif_date_col).value
                logger.info(f"Дата последней поверки текущей строки - {str_verif_date}")

                if str_verif_date not in EXC_STR:
                    year = get_verif_year_from_str(str_verif_date)
                else:
                    year = None

                if year == verif_year:
                    logger.info("Получаем номер текущего СИ")
                    temp_serial = str(worksheet.cell(row=r, column=number_col).value)
                    current_serial = temp_serial.strip()
                    logger.info(f"Заводской номер текущего СИ - {current_serial}")
                    temp_mitype = worksheet.cell(row=r, column=type_col).value
                    mitype = parse_type(temp_mitype)[0]
                    match mode:
                        case 'fgis':
                            if serial != '' and serial != current_serial:
                                print("Пропускаем этот СИ, номер не совпадает.")
                                pass
                            else:
                                print("Оппаньки, попался...берём в оборот.")
                                request_fgis(current_serial, si, verif_year, mitype, year)
                        case 'local':
                            if worksheet.cell(row=r, column=href_col).value != 'ФГИС информация о поверке':
                                current_si_type = worksheet.cell(row=r, column=type_col).value.strip() if worksheet.cell(row=r, column=type_col).value != None else worksheet.cell(row=r, column=type_col).value
                                local_request = request_local(current_serial, si, year, mitype, str_verif_date)

                                if type(local_request) == list:
                                    worksheet.cell(row=r, column=href_col).value = f"Проверить в ручном режиме. Найдено {len(local_request)} значения(й)"
                                    worksheet.cell(row=r, column=href_col).fill = FillYellow
                                    # for lst in local_request:
                                    #     if set_href(Fill, FillRed, href_col, lst, r, str_verif_date, worksheet):
                                    #         logger.info(f"Ссылка для СИ {lst.mi_number} найдена.")
                                    #         break
                                else:
                                    if set_href(Fill, FillRed, FillBlue, href_col, local_request, r, str_verif_date, worksheet):
                                        logger.info(f"Ссылка для СИ {local_request.mi_number} найдена.")
                            else:
                                worksheet.cell(row=r, column=href_col).fill = FillOrange
                        case 'change_serial':
                            continue
                else:
                    pass
    workbook.save(namefile)

def parse_type(mitype):
    lst = re.split(r'[- .,]', mitype)
    return lst

def set_href(Fill, FillRed, FillBlue, href_col, local_request, r, str_verif_date, worksheet):

    if not local_request == None:
        verif_date_current_card = datetime.strptime(local_request.verification_date, "%d.%m.%Y")
        date_current_card = verif_date_current_card.date()
        date_verif_current_row = str_verif_date.date()
        if date_verif_current_row == date_current_card:
            href = get_href(local_request)
            worksheet.cell(row=r, column=href_col).hyperlink = href
            worksheet.cell(row=r, column=href_col).value = 'ФГИС информация о поверке'
            worksheet.cell(row=r, column=href_col).style = "Hyperlink"
            worksheet.cell(row=r, column=href_col).fill = Fill
            logger.info(f"Записана ссылка в ячейку row={r}, column={href_col}")
            return True
        else:
            href = get_href(local_request)
            worksheet.cell(row=r, column=href_col).hyperlink = href
            worksheet.cell(row=r, column=href_col).value = f"ПРОВЕРИТЬ КОРРЕКТНОСТЬ, поверка во ФГИС: {verif_date_current_card}"
            worksheet.cell(row=r, column=href_col).style = "Hyperlink"
            worksheet.cell(row=r, column=href_col).fill = FillRed
            return False
    else:
        # worksheet.cell(row=r, column=href_col).value = 'Нет ссылки'
        worksheet.cell(row=r, column=href_col).fill = FillBlue
        return False

def get_index_card(lst_card, vefir_date):
    pass

def request_fgis(current_serial, si, verif_year, mitype, year):
    """
    Метод для запроса в БД ФГИС

    :param current_serial: текущий номер СИ
    :param si: наименование типа СИ
    :param verif_year: год поверки для проверки
    :param year: год поверки текущего СИ
    :return:
    """
    logger.info(
        f"Попытка запроса данных из БД ФГИС по текущему номеру СИ - {current_serial} и году поверки - {verif_year}")
    dict_request = format_dict_requests(title=si,
                                        number=current_serial,
                                        verif_year=year,
                                        mitype=mitype,
                                        rows=str(100))
    print(dict_request)
    parse_fgis.get_data_from_fgis(dict_request)

def request_local(serial, si, year, current_type, verif_date):
    """
    Обращаемся к локальной БД для получения данных по номеру, типу СИ
    и году поверки
    :param serial: номер СИ
    :param si: тип СИ
    :param year: год поверки СИ
    :param current_type: наименование типа СИ
    :return: словарь с данными по текущему СИ
    """
    print(f"Дата последней поверки текущего СИ {verif_date}, тип переменной - {type(verif_date)}")

    dict_filter = {'serial_si': serial,
                   'name_si': si,
                   'verif_year': year,
                   'type_si': current_type,
                   'verif_date': verif_date}
    db = WorkDb()
    card = db.get_card_for_si(dict_filter)
    if not type(card) == list and not card == None:
        print(card)
        return card
    else:
        if not card == None:
            print(f"Получено для текущего СИ {len(card)} значений из БД.")
        else:
            print(f"Ничего не получено для текущего СИ {serial}")
    if type(card) == list and len(card) > 0:
        d = parse_list_card(card, dict_filter['type_si'], dict_filter['verif_date'])
        if len(d) == 1:
            card = d[list(d.keys())[0]]
            return card
        else:
            return list(d.values())


def get_href(local_request):
    """
    Метод для разбора словаря с данными по СИ
    и получения ссылки из него
    :param local_request: словарь с данными
    :return: строка, содержащая ссылку на карточку СИ
    """
    return local_request.href


def parse_list_card(lst, type_si, verif_date):

    res_dict = {}
    date = datetime.strftime(verif_date, '%d.%m.%Y')
    for ind, card in enumerate(lst):
        current_type = card.mi_mitype
        current_mod = card.mi_modification
        parse_type = TypeParseSi(current_type, type_si)
        res_parse = parse_type.parse()
        parse_mod = TypeParseSi(current_mod, type_si)
        res_parse_mod = parse_mod.parse()
        if check_true(res_parse) and date == card.verification_date:
            res_dict[ind] = card
    return res_dict


def check_true(dct):
    lst = list(dct.values())
    count = 0
    for i in lst:
        if i:
            count += 1

    return True if count > 0 else False


def main():
    logger.info("Запуск скрипта!")
    arg_parser = create_parse_arg()
    namespace_arg = arg_parser.parse_args(sys.argv[1:])
    namefile = namespace_arg.namefile
    verif_year = namespace_arg.years
    keyword_si = namespace_arg.keyword
    mode = namespace_arg.mode
    serial = namespace_arg.serial
    logger.info(f"Парсим файл Excel со следующими исходными данными: имя файла - {namefile}, "
                f"год поверки - {verif_year}, СИ для парсинга - {keyword_si}")
    get_parse_si(keyword_si, verif_year=verif_year, namefile=namefile, mode=mode, serial=serial)


if __name__ == "__main__":
    main()