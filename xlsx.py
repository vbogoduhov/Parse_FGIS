#!/usr/bin/python3
"""
В данном модуле собран функционал для работы с файлом Excel
в рамках данной задачи.
"""

# Блок импорта
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Protection
import app_logger
from datetime import datetime, date
import re

# Конец блока импорта

logger = app_logger.get_logger(__name__, 'xlsx_log.log')

# Константы для заливки
GREEN = PatternFill(fill_type='solid', start_color='00FA9A', end_color='00FA9A')
RED = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
ORANGE = PatternFill(fill_type='solid', start_color='FF66FF', end_color='FF66FF')
BLUE = PatternFill(fill_type='solid', start_color='0099FF', end_color='0099FF')
YELLOW = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')

COLORS = {'green': GREEN,
          'red': RED,
          'orange': ORANGE,
          'blue': BLUE,
          'yellow': YELLOW}

COLUMNS_SI = {
    'ПУ': {
        'type': 8,
        'serial': 9,
        'verif_date': 12,
        'valid_date': 13,
        'href': 14,
        'id': 36
    },
    'ТТ': {
        'type': 15,
        'serial': 17,
        'verif_date': 20,
        'valid_date': 21,
        'href': 22,
        'id': 36
    },
    'ТН': {
        'type': 23,
        'serial': 25,
        'verif_date': 28,
        'valid_date': 29,
        'href': 30,
        'id': 36
    }
}

EXC_STR = ['',
           '-',
           '--',
           '---',
           'не пригоден',
           'н/д',
           'нет данных',
           'отсутствует',
           None]


def open_file(filename: str):
    """
    Метод для открытия файла. Возвращает объект книги.
    :param filename: имя файла.
    :return: объект workbook
    """
    try:
        workbook = openpyxl.load_workbook(filename, data_only=True)
        return workbook
    except Exception as err:
        logger.warning(f"Не удалось открыть файл {filename}. Ошибка - {err.__str__()}")


def get_worksheet(workbook: openpyxl.workbook.workbook.Workbook, namesheet: str):
    """
    Метод для получения объекта листа по имени листа
    :param workbook: объект workbook
    :return: объект worksheet
    """
    if type(workbook) == openpyxl.workbook.workbook.Workbook:
        sheet = workbook[namesheet]
        return sheet
    else:
        logger.warning(f"Невозможно получить объект sheet для типа {type(workbook)}")


def set_fill(worksheet: openpyxl.worksheet.worksheet.Worksheet, cells: tuple, color: str):
    """
    Метод для заливки ячейки цветом.
    :param worksheet: объект openpyxl.workbook.worksheet
    :param cells: кортеж с координатами ячейки (0 - строка, 1 - столбец)
    :param color: цвет заливки, заранее определён константами в модуле, ключ в словаре
    :return:
    """
    worksheet.cell(row=cells[0], column=cells[1]).fill = COLORS[color]
    logger.info(f"Ячейка с координатами ({cells[0]}, {cells[1]}) залита цветом {color}")


def unmerge(worksheet: openpyxl.worksheet.worksheet.Worksheet, coord: tuple):
    """
    Метод для снятия объединения для диапазона ячеек
    :param worksheet: объект листа
    :param coord: кортеж с координатами: 0 - start_row,
                                         1 - start_column,
                                         2 - end_row,
                                         3 - end_column
    :return:
    """
    worksheet.unmerge_cells(start_row=coord[0],
                            start_column=coord[1],
                            end_row=coord[2],
                            end_column=coord[3])


def check_merged(worksheet: openpyxl.worksheet.worksheet.Worksheet, coord: tuple):
    """
    Проверка диапазона ячеек на объединение
    :param worksheet: объект листа
    :param coord: кортеж с координатами
    :return: True or False
    """
    if type(worksheet.cell(row=coord[0], column=coord[1])) is openpyxl.cell.cell.MergedCell:
        return True
    else:
        return False


def save(workbook: openpyxl.workbook.workbook.Workbook, namefile: str):
    """
    Метод для сохранения и закрытия книги Excel
    :param workbook: объект openpyxl.workbook
    :param namefile: имя файла для сохранения
    :return: 
    """
    try:
        workbook.save(namefile)
        workbook.close()
        logger.info(f"Файл {namefile} сохранён и закрыт.")
    except Exception as err:
        logger.warning(f"Не удалось сохранить и закрыть файл {namefile}. Ошибка <{err.__str__()}>")


def merge(worksheet: openpyxl.worksheet.worksheet.Worksheet, coord: tuple):
    """
    Объединение диапазона ячеек
    :param worksheet: объект листа worksheet
    :param coord: кортеж с координатами диапазона
    :return:
    """
    try:
        worksheet.merge_cells(start_row=coord[0],
                              start_column=coord[1],
                              end_row=coord[2],
                              end_column=coord[3])
    except Exception as err:
        logger.warning(f"Не удалось объединить диапазон, ошибка <{err.__str__()}>")


# ==================================================================================== #


class XlsxFile:
    """
    Класс, представляющий струкутуру данных для файла Excel
    с параметрами и методами для работы с файлом в рамках конкретной задачи
    """
    # Константы для заливки
    # Зелёный для вновь обработанных строк, по которым есть ссылка
    GREEN = PatternFill(fill_type='solid', start_color='00FA9A', end_color='00FA9A')
    # Красный для вновь обработанных строк, по которым есть ссылка,
    # но дата последней поверки не совпадает с тем, что во ФГИС
    RED = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    # Для строк, в которых уже есть гиперссылка, она валидна, их обрабатывать не нужно
    # они пропускаются
    ORANGE = PatternFill(fill_type='solid', start_color='ffa500', end_color='ffa500')
    # Красно-коричневый, для тех строк, в которых ссылка не валидна
    RED_BROWN = PatternFill(fill_type='solid', start_color='cc8899', end_color='cc8899')
    # Голубой для строк, по которым ничего найдено не было
    BLUE = PatternFill(fill_type='solid', start_color='0099FF', end_color='0099FF')
    # Жёлтый, для строк по которым было найдено несколько значений, однозначно определиться
    # по известным критерям не получилось, поэтому нужно в ручном режиме проверить
    YELLOW = PatternFill(fill_type='solid', start_color='FFFF00', end_color='FFFF00')

    COLORS = {'green': GREEN,
              'red': RED,
              'orange': ORANGE,
              'blue': BLUE,
              'yellow': YELLOW,
              'red_brown': RED_BROWN}

    COLUMNS_SI = {
        'ПУ': {
            'type': 8,
            'serial': 9,
            'verif_date': 12,
            'valid_date': 13,
            'href': 14
        },
        'ТТ': {
            'type': 15,
            'serial': 17,
            'verif_date': 20,
            'valid_date': 21,
            'href': 22
        },
        'ТН': {
            'type': 23,
            'serial': 25,
            'verif_date': 28,
            'valid_date': 29,
            'href': 30
        }
    }

    EXC_STR = ['',
               '-',
               '--',
               '---',
               'не пригоден',
               'н/д',
               'нет данных',
               'отсутствует',
               None]

    SHEETNAME = 'Прил.1.1 (Сч,ТТ,ТН)'

    COLUMN_ID = 36

    def __init__(self, namefile: str):
        """
        Конструктор класса
        """
        self.namefile = namefile
        self.__setup()

    def __setup(self):
        """
        Инициализация файла со всеми параметрами

        :return:
        """
        self._file = openpyxl.load_workbook(self.namefile, data_only=True)
        self.active_sheet = self._file[XlsxFile.SHEETNAME]

    @property
    def max_row(self):
        return self.active_sheet.max_row

    @property
    def worksheets(self):
        return self._file.worksheets

    @property
    def mereged_range(self):
        return self.active_sheet.merged_cells

    def _read_value(self, coord: tuple):
        """
        Метод для чтения значения ячейки с указанными координатами

        :param coord: кортеж с координатами ячейки, 0 - строка, 1 - столбец
        :return: прочитанное необработанное значение
        """
        return self.active_sheet.cell(coord[0], coord[1]).value

    def _write_value(self, coord: tuple, value=None):
        """
        Метод для записи значения в ячейку с указаннми координатами

        :param coord: кортеж с координатами ячейки, 0 - строка, 1 - столбец
        :param value: само значение для записи
        :return:
        """
        try:
            self.active_sheet.cell(coord[0], coord[1]).value = value
        except Exception as err:
            logger.warning(f"Не удалось записать значение <{value}> в ячейку с координатами <{coord}>."
                           f"Ошибка: {err.__str__()}")

    def _check_date(self, date_for_check: object):
        """
        Метод для проверки даты, принадлежит ли дата к типу datetime

        :param date_for_check: дата для проверки
        :return: True или False
        """
        return True if type(date_for_check) is datetime else False

    def _change_type_date(self, date_for_change: str):
        """
        Метод для приведения даты к типу datetime, если исходный тип = str

        :param date_for_change: строка даты, которую нужно привести к типу datetime
        :return: дату типа datetime
        """
        try:
            if date_for_change not in EXC_STR:
                res_date = datetime.strptime(date_for_change, "%d.%m.%Y")
                return res_date
            else:
                return None
        except Exception as err:
            logger.warning(f"Невозможно привести дату к типу datetime. Ошибка <{err.__str__()}>")

    def _split_type(self, mitype):
        """
        Разделяем строку с типом СИ на составляющие

        :param mitype: строка типа СИ
        :return: список строк, получившийся в результате сплита
        """
        lst = re.split(r'[- .,]', mitype)
        return lst

    def _parse_type(self, lst_type):
        """
        Получаем из списка буквенное значение типа

        :param lst_type: список строк
        :return: буквенное обозначение типа СИ
        """
        tmp_val = lst_type[0]
        # for val in lst_type[1:]:
        #     if val.isalpha() and len(val) > len(tmp_val):
        #         tmp_val = tmp_val + " " + val
        return tmp_val

    def get_date(self, coord: tuple):
        """
        Метод для получения даты последней поверки из ячейки

        :param coord: кортеж с координатами ячейки
        :return: дата последней поверки в формате datetime
        """
        tmp_val = self._read_value(coord)
        if self._check_date(tmp_val):
            return tmp_val.date()
        else:
            date = self._change_type_date(tmp_val)
            return date

    def get_type(self, coord: tuple):
        """
        Метод для получения типа СИ из ячейки

        :param coord: кортеж с координатами ячейки
        :return: тип СИ
        """
        tmp_val = self._read_value(coord)
        if tmp_val is None:
            return None
        else:
            if tmp_val == 'Т-0,66' or tmp_val == 'T-0,66':
                type_si = 'Т-0,66'
            else:
                type_si = self._parse_type(self._split_type(tmp_val))
            return type_si

    def get_serial(self, coord: tuple):
        """
        Метод для получения серийного номера СИ

        :param coord: кортеж с координатами ячейки
        :return: строку серийного номера
        """
        tmp_serial = self._read_value(coord)
        if not tmp_serial is None and type(tmp_serial) is str:
            serial = tmp_serial.strip().rstrip()
            return serial
        else:
            return tmp_serial

    def get_href(self, coord: tuple):
        """
        Метод для получения ссылки из ячейки

        :param coord: кортеж с координатами ячейки
        :return: строку URL
        """
        return self.active_sheet.cell(coord[0], coord[1]).hyperlink.target

    def set_date(self, coord: tuple, date):
        """
        Метод для записи даты в ячейку

        :param coord: кортеж с координатами ячейки
        :param date: значение даты для записи
        :return:
        """
        self._write_value(coord, date)

    def get_id_record(self, coord: tuple):
        """
        Метод для получения номера идентификатора записи таблицы tbmetrology
        для текущей строки

        :param coord: кортеж с координатами ячейки
        :return: идентификатор записи
        """
        id_record = self._read_value(coord)
        if id_record is None:
            return None
        else:
            return int(id_record)

    def set_href(self, coord: tuple, href: str):
        """
        Метод для записи ссылки в ячейку

        :param coord: кортеж с координатами ячейки
        :param href: строка URL
        :return:
        """
        self._write_value(coord, "ФГИС")
        self.set_hlink_style(coord, href)
        self.set_alignment(coord, align_style="center")

    def set_hlink_style(self, coord: tuple, href: str):
        """
        Метод для присвоения формата ячейке "гиперссылка"

        :param coord: кортеж с координатами ячейки
        :return:
        """
        self.active_sheet.cell(coord[0], coord[1]).style = "Hyperlink"
        self.active_sheet.cell(coord[0], coord[1]).hyperlink = href
        logger.info(f"В ячейку с координатами <{coord}> добавлена гиперссылка {href}")

    def set_alignment(self, coord: tuple, align_style: str = "center"):
        """
        Метод для установки стиля выравнивания ячейки

        :param coord: кортеж с координатами ячейки
        :param align_style: стиль выравнивания: center
        :return:
        """
        self.active_sheet.cell(coord[0], coord[1]).alignment = Alignment(vertical=align_style, horizontal=align_style)
        logger.info(f"Для ячейки с координатами <{coord}> установлен стиль выравнивания {align_style}")

    def set_fill(self, coord: tuple, color: str):
        """
        Метод для установки заливки ячейки

        :param coord: кортеж с координатами ячейки
        :param color: цвет заливки, выбирается из предложенных заранее: green, red, blue, orange, yellow
        :return:
        """
        self.active_sheet.cell(coord[0], coord[1]).fill = XlsxFile.COLORS[color]
        logger.info(f"Для ячейки с координатами <{coord}> установлен цвет заливки {color}")

    def set_id_record(self, coord: tuple, id_record: int):
        """
        Метод для записи идентификатора из тиблицы tbmetrology

        :param coord: кортеж с координатами ячейки
        :param id_record: идентификатор записи
        :return:
        """
        current_cell = self.active_sheet.cell(coord[0], coord[1])
        current_cell.font = Font(color='00ffffff')
        current_cell.value = f"{id_record}"
        current_cell.protection = Protection(hidden=True)

    def check_merged(self, coord: tuple):
        """
        Проверка диапазона ячеек на объединение

        :param coord: кортеж с координатами
        :return: True or False
        """
        if type(self.active_sheet.cell(row=coord[0], column=coord[1])) is openpyxl.cell.cell.MergedCell:
            return True
        else:
            return False

    def unmerge(self, coord: tuple):
        """
        Метод для снятия объединения для диапазона ячеек

        :param worksheet: объект листа
        :param coord: кортеж с координатами: 0 - start_row,
                                             1 - start_column,
                                             2 - end_row,
                                             3 - end_column
        :return:
        """
        self.active_sheet.unmerge_cells(start_row=coord[0],
                                        start_column=coord[1],
                                        end_row=coord[2],
                                        end_column=coord[3])

    def merge(self, coord: tuple):
        """
        Объединение диапазона ячеек

        :param worksheet: объект листа worksheet
        :param coord: кортеж с координатами: 0 - start_row,
                                             1 - start_column,
                                             2 - end_row,
                                             3 - end_column
        :return:
        """
        try:
            self.active_sheet.merge_cells(start_row=coord[0],
                                          start_column=coord[1],
                                          end_row=coord[2],
                                          end_column=coord[3])
        except Exception as err:
            logger.warning(f"Не удалось объединить диапазон, ошибка <{err.__str__()}>")

    def check_href_style(self, coord: tuple):
        """
        Метод для проверки ячейки, установлен ли там формат гиперссылки

        :param coord: кортеж с координатами ячейки
        :return: True or False
        """
        if self.active_sheet.cell(coord[0], coord[1]).hyperlink is None:
            return False
        else:
            return True

    def get_inform_si(self, caption_si: str, row: int):
        """
        Метод для получения информации  по СИ

        :param caption_si: название типа СИ: ПУ, ТТ или ТН
        :param row: номер строки
        :return: словарь с данными по СИ
        """
        serial = self.get_serial((row, XlsxFile.COLUMNS_SI[caption_si]['serial']))
        type = self.get_type((row, XlsxFile.COLUMNS_SI[caption_si]['type']))
        verif_date = self.get_date((row, XlsxFile.COLUMNS_SI[caption_si]['verif_date']))
        valid_date = self.get_date((row, XlsxFile.COLUMNS_SI[caption_si]['valid_date']))
        flag_href_style = None
        id_record = self.get_id_record((row, XlsxFile.COLUMN_ID))

        if self.check_href_style((row, XlsxFile.COLUMNS_SI[caption_si]['href'])):
            flag_href_style = True
            href_value = self.get_href((row, XlsxFile.COLUMNS_SI[caption_si]['href']))
        else:
            flag_href_style = False
            href_value = None

        inform_si = {'serial': serial,
                     'type': type,
                     'verif_date': verif_date,
                     'valid_date': valid_date,
                     'href_value': href_value,
                     'flag_href_style': flag_href_style,
                     'id': id_record}
        return inform_si

    def get_cell(self, coord: tuple):
        """
        Метод для получения объекта ячейки.
        Для текущей задачи не нужен, но вдруг пригодится
        :param coord: кортеж с координатами ячейки
        :return: объект cell
        """
        cells = self.active_sheet.cell(coord[0], coord[1])

        return cells

    def save(self):
        """
        Метод для сохранения и закрытия файла
        :param namefile: Имя файла для сохранения
        :return:
        """
        try:
            self._file.save(self.namefile)
            self._file.close()
            logger.info(f"Файл сохранён: {namefile}")
        except Exception as err:
            logger.warning(f"Не удалось сохранить файл: {self.namefile}. Ошибка: {err.__str__()}")
